from sharepoint import SharePoint
from openpyxl import Workbook

# get clients sharepoint list
clients = SharePoint().connect_to_list(ls_name='Clients')

# create excel workbook
wb = Workbook()

dest_filepath = 'client_list.xlsx'

# create worksheet
ws = wb.active
ws.title = 'Client List'

# setting sharepoint list values to excel cells
for idx, client in enumerate(clients, 1):
    ws.cell(column=1, row=idx, value=client['Title'])
    ws.cell(column=2, row=idx, value=client['AddressInfo: Street'])
    ws.cell(column=3, row=idx, value=client['AddressInfo: City'])

# save workbook
wb.save(filename=dest_filepath)
